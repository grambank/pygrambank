import re
from collections import OrderedDict

from openpyxl import Workbook

from clldutils.misc import lazyproperty
from csvw.dsv import UnicodeWriter


class Feature(OrderedDict):
    def __init__(self, spec, wiki):
        OrderedDict.__init__(self, spec)
        self._wiki = wiki
        self.domain = OrderedDict()
        spec = self['Possible Values'].replace('multistate', '').strip()
        delimiter = ',' if ',' in spec else ';'
        for val in spec.split(delimiter):
            val, desc = re.split('\s*:\s*', val.strip())
            int(val)
            self.domain[val] = desc

    @classmethod
    def from_chunk(cls, chunk, wiki):
        items = []
        for line in chunk.split('\n'):
            if ':' in line and not line.startswith('#'):
                items.append(re.split('\s*:\s*', line.strip(), 1))
        # Make sure there are no duplicate keys:
        assert len(set(i[0] for i in items)) == len(items)
        return cls(items, wiki)

    def as_chunk(self):
        return ''.join('{0}: {1}\n'.format(k, v) for k, v in self.items())

    @lazyproperty
    def wiki(self):
        res = OrderedDict()
        title, header, lines = None, None, []
        for line in self._wiki.joinpath(
                '{0}.md'.format(self.id)).read_text(encoding='utf-8-sig').split('\n'):
            line = line.strip()
            if (not title) and line.startswith('#'):
                res['title'] = title = line.replace('#', '').strip()
            elif line.startswith('## '):
                if lines:
                    res[header] = '\n'.join(lines).strip()
                header = line.replace('## ', '').strip()
            else:
                lines.append(line)
        if lines:
            res[header] = '\n'.join(lines).strip()
        return res

    @lazyproperty
    def description(self):
        if self._wiki.joinpath('{0}.md'.format(self.id)).exists():
            return self._wiki.joinpath('{0}.md'.format(self.id)).read_text(encoding='utf-8-sig')
        return self['Clarifying Comments']

    @lazyproperty
    def id(self):
        return self['Grambank ID']

    @lazyproperty
    def name(self):
        return self['Feature']

    @lazyproperty
    def name_french(self):
        return self.get('Feature question in French')

    @lazyproperty
    def patron(self):
        return self['Feature patron']


class GB20(object):
    CHUNK_SEP = '\n\n\n\n'

    def __init__(self, path):
        self.path = path

    def iterfeatures(self, wiki):
        for chunk in self.path.read_text(encoding='utf-8').split(self.CHUNK_SEP):
            if chunk.strip():
                yield Feature.from_chunk(chunk, wiki)

    def save(self, features):
        with self.path.open('w', encoding='utf8') as fp:
            fp.write('# -*- coding: utf-8 -*-\n')
            fp.write(self.CHUNK_SEP.join(f.as_chunk() for f in features))

    def listallwiki(self, wiki):
        fn = wiki / "List-of-all-features.md"
        currentlist = []
        for m in re.finditer(
            "\*\s*(?P<gbid>(GB)?\d\d\d)\s+\[(?P<feature>[^\]]+)\]\((?P<link>[^)]+)\)",
            fn.read_text(encoding='utf-8')
        ):
            res = m.groupdict()
            if not res['gbid'].startswith('GB'):
                res['gbid'] = 'GB' + res['gbid']
            currentlist.append(res)

        active = set(f.id for f in self.iterfeatures(wiki))

        print(len(active), "active")
        print(len(currentlist) - len(active), "inactive")

        with fn.open('w', encoding='utf8') as fp:
            fp.write("Active in the GramBank feature set:\n\n")
            for spec in currentlist:
                if spec['gbid'] in active:
                    fp.write("* %(gbid)s [%(feature)s](%(link)s)\n" % spec)

            fp.write(
                "\n\nNo longer active in the GramBank feature set (for historical interest only):"
                "\n\n")
            for spec in currentlist:
                if spec['gbid'] not in active:
                    fp.write("* %(gbid)s [%(feature)s](%(link)s)\n" % spec)

    def features_sheet(self):
        cols = [
            'Grambank ID', 'Feature', 'Possible Values', 'Feature patron', 'Clarifying Comments']
        features = list(self.iterfeatures(None))
        for f in features:
            for col in f:
                if col not in cols:
                    cols.append(col)

        with UnicodeWriter(self.path.parent / 'gb20features.tsv', delimiter='\t') as writer:
            writer.writerow(cols)
            for f in features:
                writer.writerow([f.get(col, '') for col in cols])

    def grambank_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "..."
        cols = [
            'Grambank ID', 'Feature', 'Relevant unit(s)', 'Function', 'Form', 'Feature patron',
            'Clarifying Comments', 'Possible Values',
            'Value', 'Source', 'Comment']
        for i, col in enumerate(cols, start=1):
            ws.cell(column=i, row=1, value=col)

        for row, feature in enumerate(self.iterfeatures(None), start=2):
            for col, colname in enumerate(cols, start=1):
                ws.cell(column=col, row=row, value=feature.get(colname, ''))
        outdir = self.path.parent / 'For_coders'
        if not outdir.exists():
            outdir.mkdir()
        wb.save(filename=str(outdir / 'GramBank_most_updated_sheet.xlsx'))
